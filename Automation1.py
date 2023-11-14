import openpyxl as xl  # an alias for the package openpyxl
from openpyxl.chart import BarChart, Reference    # To add the bar chart to the excel file

def process_workbook (filename):
    wb = xl.load_workbook(filename)
    sheet = wb ['Sheet1']

    for row in range (2, sheet.max_row + 1):
        cell = sheet.cell (row, 3)
        corrected_price = float (cell.value) * 0.9
        corrected_price_cell = sheet.cell (row, 4)   # This is how we make a new column in excel from here
        corrected_price_cell.value = corrected_price

    values = Reference (sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)            # To select the cells we want

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save (filename)   # This one will overwrite the original file
        
